#!/usr/bin/env python3
"""
Generate ZAKO Power Personas workbook (.xlsx)

Worksheets:
  1. Component Index — exhaustive master list of all system elements
  2. def1 — Default power persona (conservative triage)
"""

import sys
sys.path.insert(0, '/Library/Developer/CommandLineTools/Library/Frameworks/Python3.framework/Versions/3.8/lib/python3.8/site-packages')

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ============================================================
# COLOR SCHEME
# ============================================================
HEADER_FILL = PatternFill("solid", fgColor="2C3E50")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
SECTION_FILL = PatternFill("solid", fgColor="34495E")
SECTION_FONT = Font(bold=True, color="FFFFFF", size=10)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

# Mode fills for def1
FULL_FILL = PatternFill("solid", fgColor="27AE60")  # green
STD_FILL = PatternFill("solid", fgColor="2ECC71")   # light green
CONS_FILL = PatternFill("solid", fgColor="F39C12")   # amber
CRIT_FILL = PatternFill("solid", fgColor="E67E22")   # orange
EMRG_FILL = PatternFill("solid", fgColor="E74C3C")   # red
ACTIVE_FONT = Font(size=9)
GATED_FONT = Font(size=9, color="999999", italic=True)


def style_header(ws, row, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP
        cell.border = THIN_BORDER


def style_data(ws, start_row, end_row, num_cols):
    for r in range(start_row, end_row + 1):
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = WRAP
            cell.border = THIN_BORDER
            cell.font = Font(size=9)


# ============================================================
# COMPONENT DATA — exhaustive system element listing
# ============================================================
# Format: (ID, Name, Layer, Sublayer, ProcessClass, Description,
#          Dependencies, DepBy, EnergyProfile, fdatasyncPerOp,
#          CPUIntensity, RAMFootprint, DiskIO, RadioReq, DisplayReq,
#          UserFacing, ZAKONative, AOSPTouchType, ImplStatus, Notes)

COMPONENTS = [
    # ── ZAKO PROTOCOL LIBRARIES ──
    ("S-24", "libzako-hash", "Submerged", "Protocol Codec", "CRITICAL",
     "BLAKE3 chain hashing: frame_hash, chain_hash, genesis_anchor, constant-time compare",
     "None", "S-10,S-11,S-14,S-20,S-21",
     "Low", "0", "2", "<1KB stack", "None", "N", "N", "N", "Y", "NONE", "Implemented",
     "Stack-allocated hasher (1,912 bytes). No heap. ~650 cycles per hash on A53."),

    ("S-30", "libzako-sign", "Submerged", "Crypto", "CRITICAL",
     "ed25519 signing: keypair gen, sign, verify, secure zero. Wraps TweetNaCl.",
     "None", "S-11,S-14,S-31,S-32",
     "Medium", "0", "3", "<2KB stack", "None", "N", "N", "N", "Y", "NONE", "Implemented",
     "~650K cycles for verify on A53. Deterministic signatures (no RNG needed for sign)."),

    ("S-31", "libzako-did", "Submerged", "Crypto", "CRITICAL",
     "DID formatter: did:key method, z6Mk prefix, BASE58BTC encode/decode",
     "S-30", "S-11",
     "Negligible", "0", "1", "<512B stack", "None", "N", "N", "N", "Y", "NONE", "Implemented",
     "Pure string formatting. No I/O."),

    ("S-20", "libzako-bitpads", "Submerged", "Protocol Codec", "CRITICAL",
     "BitPads v2.0 codec: Meta Byte, Wave/Record modes, Role A/B/C, 16 wave categories, Layer 1 session header (64-bit + CRC-15)",
     "S-24", "S-10,S-12,S-14,S-21,S-25",
     "Negligible", "0", "1", "<1KB stack", "None", "N", "N", "N", "Y", "NONE", "Implemented",
     "Pure bit manipulation. No I/O."),

    ("S-21", "libzako-bitledger", "Submerged", "Protocol Codec", "CRITICAL",
     "BitLedger v3.0 codec: Layer 2 (48-bit batch header), Layer 3 (40-bit record), conservation check, value split/join. Bitshift-optimized decode.",
     "S-20", "S-10,S-14",
     "Negligible", "0", "1", "<1KB stack", "None", "N", "N", "N", "Y", "NONE", "Implemented",
     "Manual bitshift extraction on hot path saves ~250 instructions per decode."),

    ("S-22", "libzako-c0", "Submerged", "Protocol Codec", "CRITICAL",
     "C0 Enhancement Grammar: 32 control codes, 5+3 split, P/A/C flags, 13 signal slot positions, SSP byte",
     "S-20", "S-01,S-40",
     "Negligible", "0", "1", "<256B stack", "None", "N", "N", "N", "Y", "NONE", "Implemented",
     "Single-byte encode/decode. Used for inter-daemon signaling."),

    ("S-23", "libzako-pictography", "Submerged", "Protocol Codec", "INTERACTIVE",
     "4-bit symbol codec, codebook management, Context Declaration, ALERT promotion",
     "S-22", "V-10,V-11",
     "Negligible", "0", "1", "<1KB stack", "None", "N", "Y", "N", "Y", "NONE", "Implemented",
     "Display-facing codec. Only needed when UI is active."),

    ("S-25", "libzako-padsurl", "Submerged", "Protocol Codec", "BACKGROUND",
     "pads-v1 URL codec: #1pa/ prefix, base64url encode/decode, <300 chars for SMS",
     "S-20,S-21", "S-12",
     "Negligible", "0", "1", "<512B stack", "None", "N", "N", "N", "Y", "NONE", "Implemented",
     "Only needed when sharedb is encoding for SMS transport."),

    # ── ZAKO CORE DAEMONS ──
    ("S-01", "outstack-powerd", "Submerged", "Daemon", "CRITICAL",
     "Five-mode power governor: battery/thermal sysfs polling, state machine evaluation, cgroup freezer gating, C0 signal broadcast, MODE_CHANGE/GATE/RESTORE record emission",
     "S-22,S-40,B-07", "ALL",
     "Low", "0", "1", "~64KB RSS", "Sysfs reads", "N", "N", "N", "Y", "ADDITION", "Implemented",
     "Polls every 30s. Master controller. Must never be gated. Detects charge for recovery."),

    ("S-10", "telux-ledgerd", "Submerged", "Daemon", "CRITICAL",
     "Append-only SQLite ledger: multi-chain (Islands), prepared statement cache, chain hashing, conservation enforcement, signed commits, merge records, pack compaction, fsync-before-ACK",
     "S-20,S-21,S-24,S-30,S-40", "S-12,S-14,V-10,V-11,V-12",
     "High", "1", "2", "~2MB RSS (SQLite)", "1 fdatasync/record", "N", "N", "N", "Y", "ADDITION", "Implemented",
     "Dominant energy consumer (~1,001uJ per record from fdatasync). Prepared stmts save ~5,300 instr/record. Cursor verify 13x faster."),

    ("S-11", "telux-identd", "Submerged", "Daemon", "CRITICAL",
     "Identity daemon: sovereign keypair, derived keys, DID management, capability grants (depth-limited), cascade revocation, SIGN/VERIFY/KEYGEN/GRANT/REVOKE opcodes",
     "S-30,S-31,S-40", "S-12,S-14,S-32,V-12",
     "Medium", "1", "2", "~512KB RSS", "1 fdatasync/keygen", "N", "N", "N", "Y", "ADDITION", "Implemented",
     "Key operations are CPU-intensive (ed25519). Keygen requires /dev/urandom read."),

    ("S-12", "telux-sharedb", "Submerged", "Daemon", "BACKGROUND",
     "Share daemon: 4 carriers (SMS/IP/BLE/QR), durable outbound queue (1024 entries), power-mode-aware carrier selection, retry logic (3 attempts)",
     "S-10,S-11,S-20,S-25,S-40", "V-11",
     "High", "0", "1", "~256KB RSS", "Carrier I/O", "Y", "N", "N", "Y", "ADDITION", "Implemented",
     "Queue processing is BACKGROUND. IP preferred in FULL/STD, SMS in CONS+. Queue survives restart."),

    ("S-14", "Exchange Engine", "Submerged", "Daemon", "CRITICAL",
     "Bilateral exchange: SEND/RECEIVE cycle, conservation check, merge-like atomic dual-leg posting, 32 concurrent exchanges, signature verification",
     "S-10,S-30", "V-11",
     "Medium", "0 (via S-10)", "2", "~16KB (in ledgerd)", "Via ledgerd", "N", "N", "N", "Y", "NONE", "Implemented",
     "Pure logic — embedded in ledgerd process. Can create/ack exchanges in any mode. But transmission requires sharedb (BACKGROUND)."),

    ("S-32", "Capability System", "Submerged", "Daemon", "CRITICAL",
     "GRANT/REVOKE/DELEGATE records, depth checking (max 3), cascade revocation. Part of identd.",
     "S-11", "V-12",
     "Low", "0 (via S-11)", "1", "~4KB (in identd)", "Via identd", "N", "N", "N", "Y", "NONE", "Specified",
     "Cascade revocation walks grant tree. Depth-3 max limits walk to ~64 nodes."),

    ("S-13", "telux-coverd", "Submerged", "Daemon", "INTERACTIVE",
     "Cover display daemon: time, battery, mode indicator, last ledger entry, caller ID on 1.44in ST7789 SPI panel",
     "S-01,B-04", "None",
     "Low", "0", "1", "~128KB RSS", "SPI writes", "N", "Y", "N", "Y", "ADDITION", "Specified",
     "Refresh rate can be reduced in CRIT. Frozen in EMRG — last frame persists on display."),

    ("S-40", "libzako-bus", "Submerged", "IPC", "CRITICAL",
     "System bus: Unix domain socket server, poll-based event loop, category subscription, length-prefixed framing, C0 signal routing. 16 client slots.",
     "S-22", "S-01,S-10,S-11,S-12,S-13",
     "Low", "0", "1", "~128KB RSS", "Socket I/O", "N", "N", "N", "Y", "ADDITION", "Implemented",
     "Single point of failure. Must be started before all daemons. init.rc class=core."),

    # ── ZAKO DEFERRED OPERATIONS ──
    ("Z-01", "Pack Compaction", "Submerged", "Maintenance", "DEFERRED",
     "Git-like packfile generation: concatenate sequential frame blobs, build index, NULL inline frames, preserve hashes. Part of ledgerd.",
     "S-10", "None",
     "High", "2", "3", "~8KB temp buffers", "Sequential read + transactional write", "N", "N", "N", "Y", "NONE", "Implemented",
     "CPU + I/O intensive. Only run in FULL mode when charging. Reduces long-term DB size."),

    ("Z-02", "Chain Verification Audit", "Submerged", "Maintenance", "DEFERRED",
     "Full-range cursor-based chain integrity scan. Recomputes all hashes, optionally verifies signatures.",
     "S-10,S-24,S-30", "None",
     "High", "0", "3", "~1KB per record", "Sequential read (cursor)", "N", "N", "N", "Y", "NONE", "Implemented",
     "13x faster than old per-record approach. Still O(N) — defer to FULL mode for large ranges."),

    ("Z-03", "Bundle Generation", "Submerged", "Maintenance", "DEFERRED",
     "Self-contained verifiable record bundles for device-to-device sync. ZBN1 format.",
     "S-10,S-24", "S-12",
     "Medium", "1", "2", "~4KB temp", "Read records + write bundle", "N", "N", "N", "Y", "NONE", "Specified",
     "Bundle = subset of chain with hashes for independent verification."),

    ("Z-04", "OTA Update Check", "Submerged", "Maintenance", "DEFERRED",
     "Poll distribution server for available OTA updates. Download if available.",
     "None", "None",
     "High", "0", "1", "Variable", "Network download", "Y", "N", "N", "Y", "NONE", "Specified",
     "Consumes data + battery. Strictly FULL mode only."),

    # ── ZAKO OPPORTUNISTIC OPERATIONS ──
    ("Z-10", "NLQ Engine", "Submerged", "Intelligence", "OPPORTUNISTIC",
     "Natural language query interface for ledger. Rule-based FSM, pattern matching.",
     "S-10", "V-11",
     "Medium", "0", "3", "~1MB", "Ledger reads", "N", "N", "Y", "Y", "NONE", "Specified",
     "CPU-intensive pattern matching. Pure convenience — structured UI always available."),

    ("Z-11", "Speculative Prefetch", "Submerged", "Intelligence", "OPPORTUNISTIC",
     "Pre-load counterparty data, relay polling, predictive caching.",
     "S-12", "None",
     "Medium", "0", "1", "Variable", "Network reads", "Y", "N", "N", "Y", "NONE", "Concept",
     "Pure luxury. No other component depends on it."),

    # ── KERNEL / HARDWARE ──
    ("B-01", "Kernel (Linux 4.9)", "Bedrock", "Kernel", "KERNEL",
     "Linux 4.9 CAF, MSM8937/QM215 defconfig, ZAKO patches (cgroup freezer, thermal)",
     "None", "ALL",
     "N/A", "N/A", "N/A", "N/A", "N/A", "N", "N", "N", "Y", "CONFIGURATION", "Specified",
     "Always running. Power governed by cpufreq, core parking, PM domain."),

    ("B-02", "Telux-SEC LSM", "Bedrock", "Kernel Module", "KERNEL",
     "Linux Security Module for Island boundary enforcement at kernel level",
     "B-01", "S-10",
     "Negligible", "0", "1", "Kernel space", "None", "N", "N", "N", "Y", "ADDITION", "Concept",
     "Phase 7. Supplementary to SELinux. Always loaded."),

    ("B-03", "Outstack exec gate", "Bedrock", "Kernel Module", "KERNEL",
     "LSM hook at execve() checking process class vs current power mode bitmask",
     "B-01,S-01", "ALL",
     "Negligible", "0", "1", "Kernel space", "None", "N", "N", "N", "Y", "ADDITION", "Concept",
     "Phase 7. Kernel-enforced process class gating. Faster than cgroup freezer."),

    ("B-04", "Cover display driver", "Bedrock", "Kernel Driver", "KERNEL",
     "ST7789 SPI panel init, framebuffer for 1.44in cover display",
     "B-01", "S-13",
     "Low", "0", "1", "Kernel space", "SPI writes", "N", "Y", "N", "Y", "ADDITION", "Unknown",
     "Needs DTB decompile. Panel power can be cut in EMRG."),

    ("B-05", "Lid sensor", "Bedrock", "Kernel Driver", "KERNEL",
     "Hall effect SW_LID input event, GPIO mapping. Flip close triggers CONSERVE.",
     "B-01", "S-01",
     "Negligible", "0", "1", "Kernel space", "GPIO interrupt", "N", "N", "N", "Y", "CONFIGURATION", "Specified",
     "Input event routed to outstack-powerd. Always active."),

    ("B-06", "T9 keypad driver", "Bedrock", "Kernel Driver", "KERNEL",
     "GPIO-keys scan codes for physical keypad",
     "B-01", "V-07",
     "Negligible", "0", "1", "Kernel space", "GPIO interrupt", "N", "N", "N", "Y", "CONFIGURATION", "Reference",
     "Always active. Physical keys must work even in EMRG for emergency calls."),

    ("B-07", "Power domain controller", "Bedrock", "Kernel Interface", "KERNEL",
     "Hardware power gating via sysfs/RPM: CPU freq, core online, governor, cgroups",
     "B-01", "S-01",
     "N/A", "N/A", "N/A", "Kernel space", "Sysfs writes", "N", "N", "N", "N", "CONFIGURATION", "Reference",
     "outstack-powerd HAL writes to these sysfs paths."),

    ("B-08", "dm-verity", "Bedrock", "Security", "KERNEL",
     "AVB vbmeta signing, hash tree for system partition integrity",
     "B-01", "None",
     "Low", "0", "1", "Kernel space", "Block reads", "N", "N", "N", "N", "CONFIGURATION", "Specified",
     "Always active. Verifies system partition on read. Negligible runtime cost."),

    ("B-09", "FBE (File-Based Encryption)", "Bedrock", "Security", "KERNEL",
     "File-based encryption with Keymaster 4.0 ICE (Inline Crypto Engine)",
     "B-01", "S-10,S-11",
     "Low", "0", "1", "Kernel space + ICE HW", "Inline on every I/O", "N", "N", "N", "N", "CONFIGURATION", "Specified",
     "Hardware-accelerated. Negligible power delta vs unencrypted."),

    # ── ANDROID FRAMEWORK SERVICES ──
    ("A-01", "SurfaceFlinger", "Visible", "AOSP Framework", "INTERACTIVE",
     "Display compositor. Composites all visible surfaces, drives panel refresh.",
     "B-01", "V-10,V-11,V-12,S-13,A-15",
     "High", "0", "2", "~20MB", "GPU + display controller", "N", "Y", "Y", "N", "NONE", "AOSP stock",
     "Can reduce refresh rate in CRIT. Panel can be turned off in EMRG (but Android usually handles this via screen timeout)."),

    ("A-02", "AudioFlinger", "Visible", "AOSP Framework", "INTERACTIVE",
     "Audio mixing and routing. Handles playback, recording, call audio.",
     "B-01", "A-23",
     "Medium", "0", "2", "~8MB", "Audio DSP + codec", "N", "N", "Y", "N", "NONE", "AOSP stock",
     "Call audio must work in all modes. Notification sounds can be disabled in CRIT/EMRG."),

    ("A-03", "CameraService", "Visible", "AOSP Framework", "INTERACTIVE",
     "Camera HAL interface. Photo/video capture.",
     "B-01", "V-10",
     "Very High", "0", "4", "~30MB", "ISP + sensor", "N", "Y", "Y", "N", "NONE", "AOSP stock",
     "Highest single-component power draw. Should be gated aggressively. PADS may need camera for field photos — consider BACKGROUND exception."),

    ("A-04", "LocationManager", "Visible", "AOSP Framework", "BACKGROUND",
     "GPS/GNSS, network location, fused location provider.",
     "B-01", "V-10",
     "Very High", "0", "2", "~4MB", "GNSS receiver", "Y", "N", "N", "N", "NONE", "AOSP stock",
     "GPS is ~50mW continuous. Disable in CONS+. Network location (cell tower) is nearly free — could remain available."),

    ("A-05", "WiFi Service", "Visible", "AOSP Framework", "BACKGROUND",
     "WiFi HAL, WPA supplicant, scanning, connection management.",
     "B-01", "A-21",
     "High", "0", "1", "~8MB", "WiFi radio", "Y", "N", "N", "N", "NONE", "AOSP stock",
     "WiFi radio is ~100-200mW active. Disable in CONS+. Cat S22 Flip has WiFi but it's not primary connectivity."),

    ("A-06", "Bluetooth Service", "Visible", "AOSP Framework", "BACKGROUND",
     "Bluetooth HAL, BLE GATT, classic BT profiles.",
     "B-01", "S-12",
     "Medium", "0", "1", "~4MB", "BT radio", "Y", "N", "N", "N", "BRIDGE", "AOSP stock",
     "BLE is ~10mW during scan, <1mW idle. Sharedb uses BLE for short-range frame exchange. Consider keeping BLE available in CONS for nearby exchanges."),

    ("A-07", "Sensors Service", "Visible", "AOSP Framework", "BACKGROUND",
     "Accelerometer, gyroscope, proximity, ambient light sensors.",
     "B-01", "A-01,A-17",
     "Low", "0", "1", "~2MB", "Sensor hub", "N", "N", "N", "N", "NONE", "AOSP stock",
     "Proximity sensor needed for call screen-off (CRITICAL for call power). Ambient light for auto-brightness. Accel/gyro not needed — disable in STD+."),

    ("A-08", "Vibrator Service", "Visible", "AOSP Framework", "INTERACTIVE",
     "Haptic feedback motor control.",
     "B-01", "None",
     "Low", "0", "1", "<1MB", "Motor driver", "N", "N", "N", "N", "NONE", "AOSP stock",
     "Disable in CONS+. Negligible idle cost, but each vibration is ~100mW."),

    ("A-09", "LED / Notification Light", "Visible", "AOSP Framework", "BACKGROUND",
     "Notification LED control (if present on Cat S22).",
     "B-01", "None",
     "Negligible", "0", "1", "<1MB", "GPIO", "N", "Y", "N", "N", "NONE", "AOSP stock",
     "LED is <1mW. Could remain active in all modes as a low-cost notification indicator."),

    ("A-10", "Backlight / Brightness", "Visible", "AOSP Framework", "INTERACTIVE",
     "Main display backlight control. Auto-brightness integration.",
     "B-01,A-07", "A-01",
     "Very High", "0", "1", "<1MB", "Backlight PWM", "N", "Y", "N", "N", "CONFIGURATION", "AOSP stock",
     "Backlight is 200-400mW at max. Reduce brightness threshold per mode. EMRG: minimum brightness. outstack-powerd can write to sysfs directly."),

    ("A-11", "Screen Timeout", "Visible", "AOSP Framework", "INTERACTIVE",
     "Automatic screen-off timer.",
     "B-01", "A-10",
     "N/A", "0", "1", "<1MB", "None", "N", "Y", "N", "N", "CONFIGURATION", "AOSP stock",
     "Shorten timeout per mode: FULL=2min, STD=1min, CONS=30s, CRIT=15s, EMRG=immediate (screen off unless active input)."),

    ("A-12", "Window Animations", "Visible", "AOSP Framework", "INTERACTIVE",
     "Window/transition animation scale (0x, 0.5x, 1x).",
     "A-01", "None",
     "Low", "0", "2", "~1MB", "GPU", "N", "Y", "N", "N", "CONFIGURATION", "AOSP stock",
     "Set animation_scale=0 in CONS+. Saves GPU power and perceived latency."),

    ("A-13", "Doze Mode (Android)", "Visible", "AOSP Framework", "KERNEL",
     "Android's native battery optimization: app standby, background limits, network restrictions.",
     "B-01", "ALL Android apps",
     "N/A", "N/A", "N/A", "N/A", "N/A", "N", "N", "N", "N", "CONFIGURATION", "AOSP stock",
     "Coexists with Outstack. Outstack gates ZAKO daemons; Doze gates Android apps. Disable stock battery saver to avoid conflicts (config_lowBatteryWarningLevel=0)."),

    ("A-14", "AlarmManager", "Visible", "AOSP Framework", "KERNEL",
     "Scheduled wakeups. Used by Android apps and potentially by ZAKO bridge services.",
     "B-01", "A-22,Z-04",
     "Variable", "0", "1", "<1MB", "RTC wakeup", "N", "N", "N", "N", "NONE", "AOSP stock",
     "CRITICAL alarms (outstack sampling) use wakelock. Deferred alarms (OTA check) batched into FULL mode only."),

    ("A-15", "SystemUI", "Visible", "AOSP Framework", "INTERACTIVE",
     "Status bar, notification shade, quick settings, navigation bar, lock screen.",
     "A-01", "A-16,V-08",
     "Medium", "0", "2", "~30MB", "GPU + display", "N", "Y", "Y", "N", "CONFIGURATION", "AOSP stock",
     "Always visible when screen is on. Outstack widget (V-08) lives here as a QS tile. Frozen in EMRG with all INTERACTIVE."),

    ("A-16", "Notification Manager", "Visible", "AOSP Framework", "INTERACTIVE",
     "Notification posting, channels, DND, priority handling.",
     "A-15", "S-12,V-10,V-11",
     "Low", "0", "1", "<2MB", "None", "N", "N", "Y", "N", "NONE", "AOSP stock",
     "Notifications still queue even if UI is gated — they show when INTERACTIVE restores."),

    ("A-17", "InputManager", "Visible", "AOSP Framework", "CRITICAL",
     "Input event routing: touch, keypad, hardware buttons. Dispatches to focused window.",
     "B-06", "ALL UI",
     "Low", "0", "1", "<2MB", "None", "N", "N", "Y", "N", "NONE", "AOSP stock",
     "Must run in all modes — emergency call requires keypad input even in EMRG."),

    ("A-18", "ActivityManager / LMK", "Visible", "AOSP Framework", "KERNEL",
     "Process lifecycle, OOM scoring, Low Memory Killer. Controls which processes survive memory pressure.",
     "B-01", "ALL Android apps",
     "N/A", "N/A", "N/A", "~system_server", "None", "N", "N", "N", "N", "CONFIGURATION", "AOSP stock",
     "ZAKO daemons set oom_score_adj via init.rc to survive LMK. In EMRG, LMK is most aggressive."),

    ("A-19", "Zygote / ART", "Visible", "AOSP Framework", "INTERACTIVE",
     "JVM runtime, app forking. Required for all APK-based apps.",
     "B-01", "ALL Android apps, V-20-V-23",
     "High", "0", "2", "~50MB (shared)", "None", "N", "N", "N", "N", "NONE", "AOSP stock",
     "ZAKO native apps bypass ART entirely. Zygote only needed for F-Droid apps. Could potentially be frozen in EMRG but this is radical."),

    ("A-20", "PackageManager", "Visible", "AOSP Framework", "KERNEL",
     "APK installation, permission management, intent resolution.",
     "A-19", "ALL apps",
     "Low", "0", "1", "~system_server", "None", "N", "N", "N", "N", "NONE", "AOSP stock",
     "Always available in system_server. Package installs should be deferred to FULL."),

    ("A-21", "ConnectivityService", "Visible", "AOSP Framework", "CRITICAL",
     "Network state management, default network selection, captive portal detection.",
     "B-01", "A-05,S-12",
     "Low", "0", "1", "~system_server", "None", "Y", "N", "N", "N", "CONFIGURATION", "AOSP stock",
     "Must run for modem connectivity. Captive portal pointed to babb.tel."),

    ("A-22", "JobScheduler", "Visible", "AOSP Framework", "BACKGROUND",
     "Deferred work scheduling for Android apps. Respects Doze and battery state.",
     "A-14", "V-20,V-21",
     "Variable", "0", "1", "~system_server", "Variable", "N", "N", "N", "N", "NONE", "AOSP stock",
     "Jobs deferred automatically by Doze. Outstack gate provides additional layer."),

    ("A-23", "Telephony / RIL", "Visible", "AOSP Framework", "CRITICAL",
     "Modem interface: voice calls, SMS send/receive, LTE data attach, USSD, STK (SIM Toolkit).",
     "B-01", "S-12,A-21",
     "High", "0", "1", "~10MB (rild)", "Modem radio", "Y", "N", "Y", "N", "NONE", "AOSP stock",
     "Must run in all modes. Voice calls and SMS are the final capability in EMRG. eDRX params tightened per mode by outstack-radio-helper."),

    ("A-24", "SMS Intercept", "Visible", "AOSP Bridge", "BACKGROUND",
     "BroadcastReceiver or native daemon that intercepts incoming SMS containing pads-v1 URLs and routes to telux-ledgerd via bus.",
     "A-23,S-25,S-40", "S-10",
     "Low", "0", "1", "<2MB", "None", "N", "N", "N", "Y", "BRIDGE", "Not implemented",
     "CRITICAL DESIGN DECISION: If this is BACKGROUND, incoming financial SMS not processed in CONS+. If CRITICAL, processed in all modes. Recommend: make CRITICAL."),

    ("A-25", "USB / ADB", "Visible", "AOSP Framework", "DEFERRED",
     "USB debugging, MTP file transfer, charge detection.",
     "B-01", "None",
     "Low", "0", "1", "~2MB", "USB controller", "N", "N", "N", "N", "NONE", "AOSP stock",
     "Charge detection must always work (handled by kernel, not userspace). ADB/MTP can be deferred."),

    ("A-26", "Tethering / Hotspot", "Visible", "AOSP Framework", "OPPORTUNISTIC",
     "WiFi hotspot, USB tethering, Bluetooth tethering.",
     "A-05,A-06", "None",
     "Very High", "0", "2", "~4MB", "WiFi AP mode", "Y", "N", "N", "N", "NONE", "AOSP stock",
     "Extremely power-hungry. FULL mode only."),

    ("A-27", "VPN Service", "Visible", "AOSP Framework", "BACKGROUND",
     "VPN connection management.",
     "A-21", "None",
     "Medium", "0", "1", "~4MB", "Crypto + network", "Y", "N", "N", "N", "NONE", "AOSP stock",
     "Useful but not essential. Disable in CONS+."),

    ("A-28", "Print Service", "Visible", "AOSP Framework", "OPPORTUNISTIC",
     "Network printing framework.",
     "A-05", "None",
     "Low", "0", "1", "~2MB", "None idle", "Y", "N", "N", "N", "NONE", "AOSP stock",
     "Disable in STD+. Negligible use case on Cat S22."),

    ("A-29", "Accessibility Service", "Visible", "AOSP Framework", "INTERACTIVE",
     "Screen reader, magnification, accessibility overlays.",
     "A-01,A-17", "None",
     "Low", "0", "1", "~4MB", "None", "N", "Y", "Y", "N", "NONE", "AOSP stock",
     "Must remain available when UI is active. Gated with INTERACTIVE."),

    ("A-30", "TTS Engine", "Visible", "AOSP Framework", "OPPORTUNISTIC",
     "Text-to-speech synthesis.",
     "A-02", "A-29",
     "High", "0", "3", "~10MB", "CPU + audio", "N", "N", "Y", "N", "NONE", "AOSP stock",
     "CPU-intensive synthesis. FULL mode only."),

    ("A-31", "Download Manager", "Visible", "AOSP Framework", "BACKGROUND",
     "Background file downloads.",
     "A-21", "Z-04",
     "High", "0", "1", "~2MB", "Network + eMMC", "Y", "N", "N", "N", "NONE", "AOSP stock",
     "Disable in CONS+. OTA downloads only in FULL."),

    ("A-32", "Captive Portal Detection", "Visible", "AOSP Framework", "BACKGROUND",
     "Connectivity probe (pointed to babb.tel instead of Google).",
     "A-21", "None",
     "Low", "0", "1", "~system_server", "Network probe", "Y", "N", "N", "N", "CONFIGURATION", "AOSP stock",
     "One-time on network change. Low cost. Could remain in CONS."),

    ("A-33", "NTP Time Sync", "Visible", "AOSP Framework", "BACKGROUND",
     "Network time synchronization (africa.pool.ntp.org).",
     "A-21", "None",
     "Negligible", "0", "1", "~system_server", "1 UDP packet", "Y", "N", "N", "N", "CONFIGURATION", "AOSP stock",
     "Single UDP packet. Very low cost. Could remain in CONS."),

    ("A-34", "Airplane Mode", "Visible", "AOSP Framework", "CRITICAL",
     "Radio kill switch. Disables modem, WiFi, BT simultaneously.",
     "A-23,A-05,A-06", "None",
     "N/A", "0", "1", "~system_server", "None", "Y", "N", "N", "N", "NONE", "AOSP stock",
     "User must be able to toggle airplane mode in any mode. CRITICAL."),

    ("A-35", "Keyguard / Lock Screen", "Visible", "AOSP Framework", "CRITICAL",
     "Device lock, PIN entry, emergency call access.",
     "A-17,A-15", "None",
     "Low", "0", "1", "~system_server", "Display", "N", "Y", "Y", "N", "NONE", "AOSP stock",
     "Must work in all modes. Emergency call dialer accessible from lock screen even in EMRG."),

    # ── ZAKO VISIBLE LAYER (Apps) ──
    ("V-06", "ZAKO First-Run", "Visible", "ZAKO App", "INTERACTIVE",
     "Setup wizard: language, SIM detect, PIN, privacy. Native C.",
     "V-51", "None",
     "Low", "0", "1", "~5MB", "Display", "N", "Y", "Y", "Y", "REPLACEMENT", "Specified",
     "One-time use. Not relevant to power persona except edge case (EMRG during first boot)."),

    ("V-07", "T9 IME", "Visible", "ZAKO App", "INTERACTIVE",
     "Custom input method with Bemba/Nyanja/Tonga/Lozi extended chars. Native C core.",
     "V-51,B-06", "V-10,V-11,V-12",
     "Low", "0", "1", "~3MB", "None", "N", "N", "Y", "Y", "BRIDGE", "Specified",
     "Must be available whenever INTERACTIVE apps are active. Gated together."),

    ("V-08", "Outstack Widget", "Visible", "ZAKO App", "INTERACTIVE",
     "Status bar/QS tile showing current power mode, battery, mode indicator.",
     "S-01,A-15", "None",
     "Negligible", "0", "1", "<1MB", "None", "N", "Y", "Y", "Y", "CONFIGURATION", "Specified",
     "System property read + QS tile update. Negligible cost."),

    ("V-10", "PADS App", "Visible", "ZAKO App", "INTERACTIVE",
     "Field records UI: Work Islands, forms, photos, signoffs, exchange management. Native C.",
     "S-10,S-11,V-51", "None",
     "Medium", "0 (via S-10)", "2", "~5MB", "Via ledgerd", "N", "Y", "Y", "Y", "ADDITION", "Specified",
     "Camera use for field photos is highest power draw. Consider splitting camera as separate sub-operation."),

    ("V-11", "Exchange App", "Visible", "ZAKO App", "INTERACTIVE",
     "Send/receive UI, ledger query, natural language query, exchange history. Native C.",
     "S-10,S-12,S-14,V-51", "None",
     "Medium", "0 (via S-10)", "2", "~5MB", "Via ledgerd+sharedb", "N", "Y", "Y", "Y", "ADDITION", "Specified",
     "Initiates exchanges. Requires sharedb (BACKGROUND) for actual transmission."),

    ("V-12", "Sovereignty Dashboard", "Visible", "ZAKO App", "INTERACTIVE",
     "Island management, capability grants, identity, ledger browser. Native C.",
     "S-10,S-11,S-32,V-51", "None",
     "Low", "0 (via S-10/S-11)", "1", "~5MB", "Via ledgerd+identd", "N", "Y", "Y", "Y", "ADDITION", "Specified",
     "Read-heavy. Low power cost."),

    ("V-51", "libzako-ui", "Visible", "ZAKO Toolkit", "INTERACTIVE",
     "Minimal rendering toolkit for native C apps. lvgl or NativeActivity approach.",
     "A-01", "V-06,V-07,V-10,V-11,V-12",
     "Low", "0", "1", "~2MB", "GPU/FB", "N", "Y", "Y", "Y", "BRIDGE", "Research needed",
     "Rendering approach TBD. Must support T9 keypad navigation."),

    # ── THIRD-PARTY APPS ──
    ("V-20", "F-Droid", "Visible", "Third-Party App", "DEFERRED",
     "Privileged system app for FOSS app store, silent updates.",
     "A-19,A-21", "None",
     "Medium", "0", "1", "~20MB", "Network + APK install", "Y", "Y", "Y", "N", "REPLACEMENT", "Ready",
     "App updates only in FULL/STD. Repo refresh is DEFERRED."),

    ("V-21", "ntfy + UnifiedPush", "Visible", "Third-Party App", "BACKGROUND",
     "Push notification relay replacing FCM. Maintains persistent connection.",
     "A-19,A-21", "None",
     "Medium", "0", "1", "~10MB", "Persistent TCP", "Y", "N", "N", "N", "REPLACEMENT", "Ready",
     "Persistent connection is ~5mW. Consider: keep in CONS for notification delivery, disable in CRIT+."),

    ("V-22", "Organic Maps", "Visible", "Third-Party App", "INTERACTIVE",
     "Offline OSM maps. Replaces Google Maps.",
     "A-19,A-04", "None",
     "Medium", "0", "2", "~50MB", "GPS (if navigation)", "Y", "Y", "Y", "N", "REPLACEMENT", "Ready",
     "Offline maps work without GPS/network. GPS only for live navigation."),

    ("V-23", "AOSP Dialer", "Visible", "Third-Party App", "CRITICAL",
     "Voice/SMS/USSD. Stock AOSP. Must preserve STK (SIM Toolkit).",
     "A-23,A-19", "None",
     "Low", "0", "1", "~15MB", "Modem", "Y", "Y", "Y", "N", "NONE", "AOSP stock",
     "Emergency dialer must work in EMRG. The one APK-based app that must be CRITICAL."),

    # ── HARDWARE POWER DOMAINS ──
    ("H-01", "Main Display Panel", "Bedrock", "Hardware", "INTERACTIVE",
     "480x800 IPS main display. Backlight is dominant power cost.",
     "B-01", "A-01,A-10",
     "Very High", "N/A", "N/A", "N/A", "N/A", "N", "Y", "Y", "N", "NONE", "Hardware",
     "200-400mW at max brightness. 0mW when off. Screen timeout is the biggest single power lever."),

    ("H-02", "Cover Display Panel", "Bedrock", "Hardware", "INTERACTIVE",
     "1.44in ST7789 SPI, 128x128 or 240x240. Very low power.",
     "B-04", "S-13",
     "Low", "N/A", "N/A", "N/A", "N/A", "N", "Y", "Y", "N", "NONE", "Hardware",
     "~5-10mW when on. Could remain on in CRIT as low-cost status display. Power-cut in EMRG."),

    ("H-03", "Modem/Baseband", "Bedrock", "Hardware", "CRITICAL",
     "QM215 integrated LTE modem. Voice, SMS, data.",
     "B-01", "A-23",
     "High", "N/A", "N/A", "N/A", "N/A", "Y", "N", "Y", "N", "NONE", "Hardware",
     "~100mW active, ~5mW idle (eDRX). eDRX cycle lengthened per mode. Must stay registered for emergency calls."),

    ("H-04", "WiFi Radio", "Bedrock", "Hardware", "BACKGROUND",
     "WiFi 802.11 b/g/n radio.",
     "B-01", "A-05",
     "High", "N/A", "N/A", "N/A", "N/A", "Y", "N", "N", "N", "NONE", "Hardware",
     "~100-200mW active, ~5mW idle. Disable in CONS+."),

    ("H-05", "Bluetooth Radio", "Bedrock", "Hardware", "BACKGROUND",
     "BT 4.2 with BLE support.",
     "B-01", "A-06,S-12",
     "Low", "N/A", "N/A", "N/A", "N/A", "Y", "N", "N", "N", "NONE", "Hardware",
     "BLE is <1mW idle, ~10mW scan. Consider keeping for nearby exchange in CONS."),

    ("H-06", "GPS/GNSS", "Bedrock", "Hardware", "BACKGROUND",
     "GNSS receiver for location.",
     "B-01", "A-04",
     "Very High", "N/A", "N/A", "N/A", "N/A", "Y", "N", "N", "N", "NONE", "Hardware",
     "~50mW continuous. Disable in CONS+. Cell-tower location nearly free."),

    ("H-07", "eMMC Storage", "Bedrock", "Hardware", "CRITICAL",
     "16GB eMMC flash. Primary storage. fdatasync target.",
     "B-01", "S-10,S-11",
     "High", "N/A", "N/A", "N/A", "N/A", "N", "N", "N", "N", "NONE", "Hardware",
     "~100mW during write. 2-10ms per fdatasync. Deep sleep mode in EMRG may increase latency."),

    ("H-08", "CPU (QM215 Cortex-A53)", "Bedrock", "Hardware", "CRITICAL",
     "4x Cortex-A53 @ 1.4GHz. Core parking and frequency scaling per mode.",
     "B-01,B-07", "ALL",
     "Variable", "N/A", "N/A", "N/A", "N/A", "N", "N", "N", "N", "NONE", "Hardware",
     "FULL: 4 cores perf. STD: 4 schedutil. CONS: 4 powersave. CRIT: 2 powersave. EMRG: 1 minimum."),

    ("H-09", "GPU (Adreno 308)", "Bedrock", "Hardware", "INTERACTIVE",
     "GPU for display compositing and UI rendering.",
     "B-01", "A-01,V-51",
     "Medium", "N/A", "N/A", "N/A", "N/A", "N", "Y", "N", "N", "NONE", "Hardware",
     "Idle when screen off. Active only during compositing. Can be clock-gated in CRIT."),

    ("H-10", "Charger IC / PMU", "Bedrock", "Hardware", "CRITICAL",
     "Power management unit. Charge detection, voltage regulation.",
     "B-01", "S-01",
     "N/A", "N/A", "N/A", "N/A", "N/A", "N", "N", "N", "N", "NONE", "Hardware",
     "Always active. Kernel notifies charging state via power_supply sysfs."),

    # ── INIT / INFRASTRUCTURE ──
    ("I-03", "init.rc services", "Build", "Infrastructure", "KERNEL",
     "Service entries for all ZAKO daemons. Start order, restart policy, OOM score.",
     "B-01", "S-01,S-10,S-11,S-12,S-13,S-40",
     "N/A", "N/A", "N/A", "N/A", "N/A", "N", "N", "N", "Y", "ADDITION", "Specified",
     "CRITICAL daemons: restart=always, oom_score_adj=-900. BACKGROUND: restart=on-failure."),

    ("I-04", "SELinux Policy", "Build", "Infrastructure", "KERNEL",
     "Custom domains for ZAKO daemons, file_contexts for /data/zako/.",
     "B-01", "S-01,S-10,S-11,S-12,S-13",
     "Negligible", "N/A", "N/A", "N/A", "N/A", "N", "N", "N", "Y", "ADDITION", "Specified",
     "Always enforcing. No power persona variation."),

    ("I-05", "OTA System", "Build", "Infrastructure", "DEFERRED",
     "Full + incremental OTA generation, signing, update application.",
     "A-21,A-31", "None",
     "High", "0", "1", "Variable", "Large download + flash", "Y", "N", "N", "Y", "NONE", "Specified",
     "Strictly FULL mode, charging preferred. Large power draw during apply."),

    # ── SPECIAL / CROSS-CUTTING ──
    ("X-01", "Wakelock Management", "Visible", "Cross-cutting", "CRITICAL",
     "ZAKO daemons hold partial wakelocks during CRITICAL operations to prevent suspend.",
     "B-01", "S-01,S-10,S-11",
     "Variable", "N/A", "N/A", "N/A", "N/A", "N", "N", "N", "Y", "NONE", "Design needed",
     "CRITICAL daemons hold wakelocks during record append (fdatasync must complete). Release immediately after."),

    ("X-02", "eDRX Radio Parameters", "Visible", "Cross-cutting", "CRITICAL",
     "Extended discontinuous reception parameters tightened per power mode.",
     "A-23,S-01", "H-03",
     "N/A", "N/A", "N/A", "N/A", "N/A", "Y", "N", "N", "Y", "BRIDGE", "Specified",
     "FULL: eDRX off. STD: 5.12s. CONS: 20.48s. CRIT: 40.96s. EMRG: 81.92s. Affects SMS delivery latency."),

    ("X-03", "Cgroup Freezer Hierarchy", "Visible", "Cross-cutting", "CRITICAL",
     "cgroup freezer groups for each process class: /sys/fs/cgroup/freezer/outstack_{opp,def,bg,int}/",
     "B-01,S-01", "ALL gated processes",
     "Negligible", "N/A", "N/A", "N/A", "Sysfs writes", "N", "N", "N", "Y", "ADDITION", "Specified",
     "outstack-powerd writes FROZEN/THAWED to freezer.state files. Memory-backed (procfs), not eMMC."),

    ("X-04", "Emergency Call Path", "Visible", "Cross-cutting", "CRITICAL",
     "Full path from keypad → InputManager → Dialer → RIL → modem for emergency calls.",
     "B-06,A-17,A-23,A-35,V-23", "None",
     "N/A", "N/A", "N/A", "N/A", "N/A", "Y", "Y", "Y", "N", "NONE", "Design needed",
     "Must work in ALL modes including EMRG. Every component in this chain must be CRITICAL or KERNEL. This is a regulatory requirement."),

    ("X-05", "Charging Recovery Path", "Visible", "Cross-cutting", "CRITICAL",
     "Detection of charge plug-in → battery rise → mode transitions upward through hysteresis.",
     "H-10,S-01", "ALL",
     "N/A", "N/A", "N/A", "N/A", "N/A", "N", "N", "N", "Y", "NONE", "Implemented",
     "Recovery: EMRG→CRIT at 6%, CRIT→CONS at 13%, CONS→STD at 23%, STD→FULL at 53%. Charge override skips modes when fast-charging above 30%."),
]

# Column headers for Component Index
INDEX_HEADERS = [
    "ID", "Component", "Layer", "Sublayer", "Process Class",
    "Description", "Dependencies", "Depended-On-By",
    "Energy Profile", "fdatasync/Op", "CPU Intensity (1-5)",
    "RAM Footprint", "Disk I/O", "Radio Required",
    "Display Required", "User-Facing", "ZAKO Native",
    "AOSP Touch Type", "Impl Status", "Notes"
]

INDEX_WIDTHS = [
    8, 22, 12, 16, 16,
    50, 24, 24,
    12, 10, 12,
    14, 18, 10,
    10, 10, 10,
    14, 14, 50
]

# ============================================================
# SHEET 1: Component Index
# ============================================================
ws_index = wb.active
ws_index.title = "Component Index"

# Write headers
for col, header in enumerate(INDEX_HEADERS, 1):
    ws_index.cell(row=1, column=col, value=header)
style_header(ws_index, 1, len(INDEX_HEADERS))

# Write data
for row_idx, comp in enumerate(COMPONENTS, 2):
    for col_idx, val in enumerate(comp, 1):
        ws_index.cell(row=row_idx, column=col_idx, value=val)

style_data(ws_index, 2, len(COMPONENTS) + 1, len(INDEX_HEADERS))

# Set column widths
for col, width in enumerate(INDEX_WIDTHS, 1):
    ws_index.column_dimensions[get_column_letter(col)].width = width

# Freeze top row
ws_index.freeze_panes = "A2"
# Auto-filter
ws_index.auto_filter.ref = f"A1:{get_column_letter(len(INDEX_HEADERS))}{len(COMPONENTS)+1}"


# ============================================================
# DEF1 PERSONA DATA
# ============================================================
# Format: (ID, Component, ProcessClass,
#          FULL_behavior, STD_behavior, CONS_behavior, CRIT_behavior, EMRG_behavior,
#          CascadeRisk, RecoveryPriority)

DEF1 = [
    # CRITICAL class — active in all modes
    ("S-01", "outstack-powerd", "CRITICAL",
     "Polls 30s. All cores perf. Full evaluation.", "Same", "Same", "Same, 2 cores", "Same, 1 core min freq. Charge detection active.",
     "SINGLE POINT: if this dies, no mode recovery. init.rc restarts.", "1 (always running)"),

    ("S-40", "libzako-bus", "CRITICAL",
     "Full IPC routing. 16 client slots.", "Same", "Same", "Same", "Same. Fewer active clients.",
     "SINGLE POINT: all daemon IPC depends on bus. init.rc restart=always.", "1 (always running)"),

    ("S-10", "telux-ledgerd", "CRITICAL",
     "Full: append, verify, batch, pack. Prepared stmt cache active.", "Same", "Same, no inbound from sharedb", "Same, slower (2 cores, powersave)", "Listening but no frames arriving. Chain intact.",
     "No cascade: pure append-only. Idle if no input.", "1 (always running)"),

    ("S-11", "telux-identd", "CRITICAL",
     "Full: sign, verify, keygen, grant, revoke.", "Same", "Same", "Same, slower keygen", "Listening but no requests arriving.",
     "No cascade: stateless request/response.", "1 (always running)"),

    ("S-14", "Exchange Engine", "CRITICAL",
     "Full bilateral exchange. Merge-like atomic posting.", "Same", "Can create/ack locally. ACKs don't transmit (sharedb frozen).", "Same as CONS, slower", "No exchanges possible (no UI, no sharedb).",
     "Exchange completion stalls in CONS+ — ACKs queue in sharedb.", "1 (always running, part of ledgerd)"),

    ("A-17", "InputManager", "CRITICAL",
     "Full input routing.", "Same", "Same", "Same", "Keypad input only (for emergency calls).",
     "Must work for emergency call path.", "1 (always running)"),

    ("A-21", "ConnectivityService", "CRITICAL",
     "Full network management.", "Same", "Same", "Same", "Minimal — modem registration only.",
     "Modem registration depends on this.", "1 (always running)"),

    ("A-23", "Telephony / RIL", "CRITICAL",
     "Full: voice, SMS, data, USSD, STK.", "Same. eDRX 5.12s.", "Same. eDRX 20.48s. IP transport deprioritized.", "Same. eDRX 40.96s. SMS delayed 20-40s.", "Voice + SMS only. eDRX 81.92s. Data detached.",
     "Emergency calls require full RIL path.", "1 (always running)"),

    ("V-23", "AOSP Dialer", "CRITICAL",
     "Full dialer/contacts/SMS.", "Same", "Same", "Same", "Emergency dialer only (from lock screen).",
     "Emergency call path: keypad→dialer→RIL→modem.", "1 (always running)"),

    ("A-35", "Keyguard", "CRITICAL",
     "Full lock screen.", "Same", "Same", "Same", "Emergency call access from lock screen.",
     "Emergency call access.", "1 (always running)"),

    ("X-04", "Emergency Call Path", "CRITICAL",
     "Full path active.", "Same", "Same", "Same", "MUST WORK. Regulatory requirement.",
     "Every component in chain must be CRITICAL.", "1 (always running)"),

    # INTERACTIVE class — gated in EMRG
    ("A-01", "SurfaceFlinger", "INTERACTIVE",
     "Full compositing, normal refresh.", "Same", "Same", "Reduced refresh possible", "GATED. Screen off or frozen.",
     "All UI depends on SurfaceFlinger.", "2 (restore at CRIT, 6% charge)"),

    ("A-02", "AudioFlinger", "INTERACTIVE",
     "Full audio. Notifications.", "Same", "Same", "No notification sounds", "GATED. No audio except call audio (handled by modem DSP).",
     "Call audio bypasses AudioFlinger via modem DSP in EMRG.", "2"),

    ("A-08", "Vibrator", "INTERACTIVE",
     "Full haptics.", "Same", "Disabled", "Disabled", "GATED.",
     "None.", "5 (lowest priority)"),

    ("A-15", "SystemUI", "INTERACTIVE",
     "Full status bar, QS tiles, notifications.", "Same", "Same", "Same, simplified", "GATED. No status bar.",
     "Outstack widget (V-08) frozen. User loses mode indicator.", "2"),

    ("A-19", "Zygote / ART", "INTERACTIVE",
     "Full JVM for APK apps.", "Same", "Same", "Same", "GATED. No APK apps can launch.",
     "All third-party apps (F-Droid, Organic Maps) require Zygote.", "2"),

    ("S-13", "telux-coverd", "INTERACTIVE",
     "Full cover updates: time, battery, mode, last entry.", "Same", "Same", "Reduced refresh (every 60s)", "GATED. Last frame persists on cover panel.",
     "Informational only. Consider: make this CRITICAL for low-cost status in EMRG?", "2"),

    ("H-01", "Main Display", "INTERACTIVE",
     "Full brightness, 2min timeout.", "Same", "1min timeout, reduced brightness", "30s timeout, minimum brightness", "GATED. Screen off unless keypad input.",
     "Biggest power lever. Each second of screen-on costs ~300mW.", "2"),

    ("H-09", "GPU", "INTERACTIVE",
     "Full compositing speed.", "Same", "Same", "Clock-gated", "GATED.",
     "No display compositing needed when screen off.", "2"),

    ("V-10", "PADS App", "INTERACTIVE",
     "Full: forms, photos, field records.", "Same", "Same (but camera separately governed)", "Same", "GATED. No field recording.",
     "Field data queued locally. No data loss.", "2"),

    ("V-11", "Exchange App", "INTERACTIVE",
     "Full: send, receive, query, NLQ.", "Same (no NLQ)", "Same (no NLQ, no outbound delivery)", "Same", "GATED. No user-initiated exchanges.",
     "Exchange initiation blocked. Existing PENDING exchanges preserved.", "2"),

    ("V-12", "Sovereignty Dashboard", "INTERACTIVE",
     "Full: islands, grants, ledger browse.", "Same", "Same", "Same", "GATED.",
     "None. Read-only.", "4"),

    ("S-23", "libzako-pictography", "INTERACTIVE",
     "Full symbol codec.", "Same", "Same", "Same", "GATED (no UI to display).",
     "None.", "4"),

    # BACKGROUND class — gated in CONS+
    ("S-12", "telux-sharedb", "BACKGROUND",
     "Full queue processing. IP preferred. All carriers.", "Same", "GATED. Queue frozen. No outbound delivery.", "GATED.", "GATED.",
     "CRITICAL CASCADE: exchange ACKs don't reach counterparty. Outbound queue durable — drains on recovery.", "3 (restore at CONS, 13% charge)"),

    ("A-24", "SMS Intercept", "BACKGROUND",
     "Intercepts pads-v1 SMS, routes to ledgerd.", "Same", "GATED. Incoming financial SMS stored but not processed.", "GATED.", "GATED.",
     "DESIGN DECISION: consider making CRITICAL to process inbound transactions in all modes.", "3"),

    ("A-04", "LocationManager", "BACKGROUND",
     "GPS available. Fused location.", "Same", "GATED. No GPS. Cell-tower location only.", "GATED.", "GATED.",
     "PADS field recording loses GPS coords. Cell-tower fallback available if made CRITICAL.", "3"),

    ("A-05", "WiFi", "BACKGROUND",
     "Full WiFi.", "Same", "GATED. WiFi off.", "GATED.", "GATED.",
     "LTE data unaffected (modem is CRITICAL).", "3"),

    ("A-06", "Bluetooth", "BACKGROUND",
     "Full BT + BLE.", "Same", "GATED. No BLE exchange.", "GATED.", "GATED.",
     "Sharedb loses BLE carrier. Consider: keep BLE in CONS for nearby exchange?", "3"),

    ("A-07", "Sensors", "BACKGROUND",
     "All sensors.", "Accel/gyro disabled. Prox + ALS remain.", "GATED.", "GATED.", "GATED.",
     "Proximity sensor needed for call screen-off. Consider: make proximity CRITICAL.", "3"),

    ("A-22", "JobScheduler", "BACKGROUND",
     "Jobs execute per schedule.", "Same", "GATED. Jobs deferred.", "GATED.", "GATED.",
     "Android app background work deferred. No ZAKO dependency.", "3"),

    ("A-27", "VPN", "BACKGROUND",
     "Active if configured.", "Same", "GATED.", "GATED.", "GATED.",
     "Network traffic unencrypted in CONS+ if VPN was active.", "3"),

    ("A-31", "Download Manager", "BACKGROUND",
     "Downloads active.", "Same", "GATED. Downloads paused.", "GATED.", "GATED.",
     "Incomplete downloads resume on recovery.", "3"),

    ("A-32", "Captive Portal", "BACKGROUND",
     "Probes on network change.", "Same", "GATED. No probe.", "GATED.", "GATED.",
     "Network may show '!' icon. Cosmetic only.", "5"),

    ("A-33", "NTP Sync", "BACKGROUND",
     "Periodic sync.", "Same", "GATED. Clock drifts.", "GATED.", "GATED.",
     "Modem provides network time (NITZ) as fallback. CRITICAL.", "5"),

    ("V-21", "ntfy/UnifiedPush", "BACKGROUND",
     "Persistent push connection.", "Same", "GATED. No push notifications.", "GATED.", "GATED.",
     "Notifications queue server-side, deliver on recovery.", "3"),

    ("H-04", "WiFi Radio", "BACKGROUND",
     "On.", "On.", "GATED. Radio off.", "Off.", "Off.",
     "None.", "3"),

    ("H-05", "BT Radio", "BACKGROUND",
     "On.", "On.", "GATED. Radio off.", "Off.", "Off.",
     "BLE exchange unavailable.", "3"),

    ("H-06", "GPS/GNSS", "BACKGROUND",
     "Available.", "Available.", "GATED. Receiver off.", "Off.", "Off.",
     "Location falls back to cell tower.", "3"),

    # DEFERRED class — gated in CONS+
    ("Z-01", "Pack Compaction", "DEFERRED",
     "Runs when idle + charging.", "Deferred.", "GATED.", "GATED.", "GATED.",
     "DB grows without compaction. No functional impact.", "4 (restore at STD, 23%)"),

    ("Z-02", "Chain Verification", "DEFERRED",
     "Scheduled audit of full chains.", "Deferred.", "GATED.", "GATED.", "GATED.",
     "Implicit per-record verification still active in lds_append.", "4"),

    ("Z-03", "Bundle Generation", "DEFERRED",
     "Creates sync bundles.", "Deferred.", "GATED.", "GATED.", "GATED.",
     "QR sync shows single frame instead of bundle.", "4"),

    ("Z-04", "OTA Update", "DEFERRED",
     "Check + download + apply.", "GATED.", "GATED.", "GATED.", "GATED.",
     "Device stays on current version.", "4"),

    ("V-20", "F-Droid", "DEFERRED",
     "Repo refresh + updates.", "Updates only.", "GATED.", "GATED.", "GATED.",
     "Apps stay at current version.", "4"),

    ("A-25", "USB/ADB", "DEFERRED",
     "Full ADB + MTP.", "ADB only.", "GATED.", "GATED.", "GATED.",
     "Charge detection always works (kernel). Debug inaccessible.", "4"),

    ("I-05", "OTA System", "DEFERRED",
     "Full OTA pipeline.", "Deferred.", "GATED.", "GATED.", "GATED.",
     "None.", "4"),

    # OPPORTUNISTIC class — gated in STD+
    ("Z-10", "NLQ Engine", "OPPORTUNISTIC",
     "Natural language queries available.", "GATED.", "GATED.", "GATED.", "GATED.",
     "Zero dependency. Structured UI always available.", "5 (restore at FULL, 53%)"),

    ("Z-11", "Speculative Prefetch", "OPPORTUNISTIC",
     "Predictive caching active.", "GATED.", "GATED.", "GATED.", "GATED.",
     "Zero dependency.", "5"),

    ("A-26", "Tethering/Hotspot", "OPPORTUNISTIC",
     "Available.", "GATED.", "GATED.", "GATED.", "GATED.",
     "None.", "5"),

    ("A-28", "Print Service", "OPPORTUNISTIC",
     "Available.", "GATED.", "GATED.", "GATED.", "GATED.",
     "None.", "5"),

    ("A-30", "TTS Engine", "OPPORTUNISTIC",
     "Available.", "GATED.", "GATED.", "GATED.", "GATED.",
     "Accessibility impact if TTS user. Consider: make INTERACTIVE.", "5"),
]

DEF1_HEADERS = [
    "ID", "Component", "Process Class",
    "FULL (100-50%)", "STD (50-20%)", "CONS (20-10%)",
    "CRIT (10-3%)", "EMRG (3-0%)",
    "Cascade Risk", "Recovery Priority"
]

DEF1_WIDTHS = [8, 22, 14, 40, 30, 40, 40, 40, 50, 18]

MODE_FILLS = {3: FULL_FILL, 4: STD_FILL, 5: CONS_FILL, 6: CRIT_FILL, 7: EMRG_FILL}

# ============================================================
# SHEET 2: def1 persona
# ============================================================
ws_def1 = wb.create_sheet("def1")

for col, header in enumerate(DEF1_HEADERS, 1):
    ws_def1.cell(row=1, column=col, value=header)
style_header(ws_def1, 1, len(DEF1_HEADERS))

# Mode column header fills
for col, fill in MODE_FILLS.items():
    cell = ws_def1.cell(row=1, column=col)
    cell.fill = fill

for row_idx, entry in enumerate(DEF1, 2):
    for col_idx, val in enumerate(entry, 1):
        cell = ws_def1.cell(row=row_idx, column=col_idx, value=val)
        # Color mode columns based on whether component is gated
        if col_idx in MODE_FILLS:
            cell.fill = MODE_FILLS[col_idx]
            if val and "GATED" in str(val):
                cell.font = GATED_FONT
            else:
                cell.font = ACTIVE_FONT

style_data(ws_def1, 2, len(DEF1) + 1, len(DEF1_HEADERS))

for col, width in enumerate(DEF1_WIDTHS, 1):
    ws_def1.column_dimensions[get_column_letter(col)].width = width

ws_def1.freeze_panes = "A2"
ws_def1.auto_filter.ref = f"A1:{get_column_letter(len(DEF1_HEADERS))}{len(DEF1)+1}"

# ============================================================
# SHEET 3: Persona Index (listing of all personas)
# ============================================================
ws_personas = wb.create_sheet("Persona Index")

PERSONA_HEADERS = ["Persona ID", "Name", "Description", "Priority Focus",
                   "Target Use Case", "Sheet Name", "Status", "Notes"]
PERSONA_WIDTHS = [12, 20, 50, 30, 30, 14, 12, 50]

for col, header in enumerate(PERSONA_HEADERS, 1):
    ws_personas.cell(row=1, column=col, value=header)
style_header(ws_personas, 1, len(PERSONA_HEADERS))

PERSONAS = [
    ("def1", "Default Conservative",
     "Standard 5-mode triage. OPPORTUNISTIC→DEFERRED→BACKGROUND→INTERACTIVE shed in order. Emergency = voice+SMS only.",
     "Battery longevity, data preservation",
     "General daily use, Zambian field conditions",
     "def1", "Active",
     "Baseline persona. All others are variations on gating order, thresholds, or class assignments."),
    ("field1", "Field Worker",
     "PADS-optimized. Camera kept available longer. GPS kept in CONS. Exchange app prioritized.",
     "Field data collection continuity",
     "Agricultural extension workers, health workers",
     "(TBD)", "Planned",
     "May reclassify camera as BACKGROUND, GPS as CRITICAL (cell-tower). PADS app sub-operations split by power cost."),
    ("trade1", "Market Trader",
     "Exchange-optimized. Sharedb kept active longer. SMS intercept CRITICAL. Short screen timeouts.",
     "Transaction completion, counterparty reliability",
     "Market vendors doing high-volume small exchanges",
     "(TBD)", "Planned",
     "Key change: SMS intercept → CRITICAL. Sharedb → INTERACTIVE (frozen only in EMRG). BLE kept for nearby exchange."),
    ("night1", "Overnight / Idle",
     "Aggressive conservation. Screen always off. Only modem + ledgerd + identd. All UI frozen.",
     "Maximum standby time",
     "Device sitting overnight, not actively used",
     "(TBD)", "Planned",
     "Could achieve <10mW idle. Start with STD-equivalent gating even at 100% battery."),
    ("emrg1", "Extended Emergency",
     "Stay alive as long as possible. Only emergency calls + modem registration. Everything else frozen.",
     "Maximum time to emergency call capability",
     "Lost in field, no charger, need phone to last days",
     "(TBD)", "Planned",
     "Even CRITICAL ledgerd/identd could be frozen if not needed. Only modem + keypad + dialer."),
    ("charge1", "Charging / Plugged In",
     "Aggressive restoration. Run all deferred maintenance (pack, verify, OTA) while charging.",
     "Maintenance completion, data hygiene",
     "Device plugged in at home/office",
     "(TBD)", "Planned",
     "Override: if charging && rate>10W, run FULL even at 10% battery. Run pack compaction, chain audits, OTA."),
    ("demo1", "Demonstration / Showroom",
     "All features active regardless of battery. No conservation. For demos and testing.",
     "Feature visibility",
     "Investor demos, trade shows, testing",
     "(TBD)", "Planned",
     "Force FULL mode. Disable all gating. Accept battery drain."),
    ("secure1", "High Security",
     "Frequent chain verification. All records signed. No deferral of crypto operations.",
     "Tamper evidence, verification freshness",
     "High-value transactions, audit scenarios",
     "(TBD)", "Planned",
     "Chain verification audit → CRITICAL. All appends require signature. Higher power cost accepted."),
]

for row_idx, persona in enumerate(PERSONAS, 2):
    for col_idx, val in enumerate(persona, 1):
        ws_personas.cell(row=row_idx, column=col_idx, value=val)

style_data(ws_personas, 2, len(PERSONAS) + 1, len(PERSONA_HEADERS))

for col, width in enumerate(PERSONA_WIDTHS, 1):
    ws_personas.column_dimensions[get_column_letter(col)].width = width

ws_personas.freeze_panes = "A2"

# ============================================================
# SAVE
# ============================================================
outpath = "/Users/mp/Documents/Vaults/babb/mobileos/zako-os-v1/system/power-personas.xlsx"
wb.save(outpath)
print(f"Written: {outpath}")
print(f"  Component Index: {len(COMPONENTS)} elements")
print(f"  def1 persona: {len(DEF1)} entries")
print(f"  Persona Index: {len(PERSONAS)} personas defined")
